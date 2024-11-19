import os
import csv
import pandas as pd
from datetime import datetime
import openpyxl
from functools import cmp_to_key
from config import *

class Order:
    def __init__(self, sku, itemDescription, itemPrice, orderNumber, orderTotal, paidByCustomer, tax, itemQty, qtyInEach, shipping):
        self.sku = sku
        self.itemDescription = itemDescription
        self.itemPrice = float(itemPrice)
        self.orderNumber = orderNumber
        self.totalSaleOrder = float(itemPrice) * int(itemQty)
        self.orderTotal = float(orderTotal)
        self.paidByCustomer = float(paidByCustomer)
        self.tax = float(tax)
        self.itemQty = int(itemQty)
        self.qtyInEach = int(qtyInEach)
        self.shipping = float(shipping)
    
    def __str__(self):
        return 'sku: {}, itemPrice: {}, qty: {}, qtyInEach: {}'.format(self.sku, self.itemPrice, self.itemQty, self.qtyInEach)

def getTimestamp():
    now = datetime.now()
    return datetime.strftime(now, "%m%d%Y%H%M%S")

def getCurrentime():
    return datetime.now()

def getFileModifiedDate(filepath):
    return datetime.fromtimestamp(os.path.getmtime(filepath))

def getDaysDifferent(currentTime, timestamp):
    return (currentTime - timestamp).days

def roundCurrency(cur):
    return round(cur, 4)

def sortOrders(a, b):
    if a[2] == 'CASE' and (b[2] == 'BOX' or b[2] == 'EA'):
        return -1
    elif a[2] == 'BOX' and b[2] == 'EA':
        return -1
    else:
        return 1

def getUOMMasterData(inputFilepath):
    mapped = {}
    message = None

    try:
        age = getDaysDifferent(getCurrentime(), getFileModifiedDate(inputFilepath))
        message = 'UOM master file was updated {} days ago.'.format(age)

        with open (inputFilepath, mode='r') as file:
            content = csv.reader(file)
            for line in content:
                if (len(line) == 3):
                    mapped[line[0]] = {
                        'item_number': line[1],
                        'uom': line[2]
                    }
    except:
        print('*** Error: Failed to read input file for UOM Master. Please make sure filename is valid. ***')
        return {}, message

    return mapped, message

def getInventoryMasterData(inputFilepath):
    mapped = {}
    message = None
    
    try:
        age = getDaysDifferent(getCurrentime(), getFileModifiedDate(inputFilepath))
        message = 'Inventory master file was updated {} days ago.'.format(age)

        workbook = openpyxl.load_workbook(inputFilepath) # #, Item No., Item Desc., Available Qty
        sheet = workbook.active
        for r in range(2, sheet.max_row+1):
            itemNumber = None
            for c in range(1, sheet.max_column+1):
                data = sheet.cell(row=r, column=c).value
                if c == 2:
                    itemNumber = str(data)
                if c == 4:
                    mapped[itemNumber] = data

        # with open (inputFilepath, mode='r') as file:
        #     content = csv.reader(file)
        #     for line in content:
        #         if (len(line) == 4):
        #             mapped[line[1]] = int(line[3]) if line[3].isnumeric() else 0
    except:
        print('*** Error: Failed to read input file for Inventory Master. Please make sure filename is valid. ***')
        return {}, message

    return mapped, message

def getOrdersFromInputfile(filepath, uomMaster):
    orders = []
    itemNumbersNotInUOMMaster = []
    
    try:
        with open (filepath, mode='r') as file:
            count = 1
            content = csv.reader(file)
            for line in content:
                if count == 1:
                    count += 1
                    continue
        
                if (len(line) == 8):
                    uomQty = int(uomMaster[line[0]]['uom']) if line[0] in uomMaster else None
                    if not uomQty:
                        # message = 'Item number <{}> is not in the UOM master data.'.format(line[0])
                        # return [], message
                        itemNumbersNotInUOMMaster.append(line[0])
                        continue
                    qtyInEach = uomQty * int(line[6])
                    order = Order(line[0], '', line[1], line[2], line[3], line[4], line[5], line[6], qtyInEach, line[7])
                    orders.append(order)
                count += 1
    except Exception as err:
        message = 'Please check your input batch file: {}'.format(filepath)
        print('*** Error: Failed to read batch input file. Please make sure filename is valid. err: {} ***'.format(err))
        return -1, [], message

    if itemNumbersNotInUOMMaster:
        message = 'One or more item numbers are not in the UOM master data.'
        return 0, itemNumbersNotInUOMMaster, message

    return 1, orders, ''

def combineOrders(orders, uomMaster):
    groupedByOrderNumber = {}
    groupedBySKU = {}

    processedOrders = {}
    orderDetails = {
        'totalOrderAmount': 0.0,
        'totalPaidByCustomer': 0.0,
        'totalOrderTax': 0.0,
        'totalShipping': 0.0
    }

    for order in orders:
        # group by Order Number
        if order.orderNumber not in groupedByOrderNumber:
            groupedByOrderNumber[order.orderNumber] = {
                'orderTotal': order.orderTotal,
                'orderPaidByCustomer': order.paidByCustomer,
                'orderTax': order.tax,
                'orderShipping': order.shipping
            }

        # group by SKU
        actualSku = uomMaster[order.sku]['item_number']
        if actualSku in groupedBySKU:
            groupedBySKU[actualSku].append(order)
        else:
            groupedBySKU[actualSku] = [order]

    for sku, orders in groupedBySKU.items():
        totalQty = 0
        totalPrice = 0
        for order in orders:
            totalQty += order.qtyInEach
            totalPrice += order.totalSaleOrder
        processedOrders[sku] = {
            'sku': sku,
            'desc': order.itemDescription,
            'totalQty': int(totalQty),
            'totalPrice': float(totalPrice),
            'pricePerPiece': float(totalPrice) / int(totalQty)
        }

    for orderNumber, order in groupedByOrderNumber.items():
        orderDetails['totalOrderAmount'] += order['orderTotal']
        orderDetails['totalPaidByCustomer'] += order['orderPaidByCustomer']
        orderDetails['totalOrderTax'] += order['orderTax']
        orderDetails['totalShipping'] += order['orderShipping']

    return processedOrders, orderDetails

def getOrdersWithUOMVariants(results, order, caseQty, boxQty):
    caseQty = int(caseQty)
    boxQty = int(boxQty)
    subTotal = 0
    caseNumber = 0
    boxNumber = 0
    eachNumber = 0
    totalQty = order['totalQty']
    itemDesc = order['desc']
    pricePerPiece = order['pricePerPiece']
 
    if caseQty != 0:
        caseNumber = totalQty // caseQty
    if boxQty != 0:
        boxNumber = (totalQty - (caseNumber * caseQty)) // boxQty
    eachNumber = totalQty - (caseNumber * caseQty) - (boxNumber * boxQty)
    
    if caseNumber > 0:
        results.append([
            order['sku'],
            itemDesc,
            'CASE',
            caseNumber,
            pricePerPiece
        ])
        subTotal += pricePerPiece * caseQty * caseNumber
    if boxNumber > 0:
        results.append([
            order['sku'],
            itemDesc,
            'BOX',
            boxNumber,
            pricePerPiece
        ])
        subTotal += pricePerPiece * boxQty * boxNumber
    if eachNumber > 0:
        results.append([
            order['sku'],
            itemDesc,
            'EA',
            eachNumber,
            pricePerPiece
        ])
        subTotal += pricePerPiece * eachNumber
    
    return subTotal

def resultIsValidated(resultDetails, orders, inventoryMaster):
    outOfStockSKUs = []
    cannotFindStockSKUs = []
    isOutOfStock = False

    if roundCurrency(resultDetails['grandTotal']) != roundCurrency(resultDetails['totalOrderBeforeDiscount']):
        return False, 'Total Before Discount does not match with Grand Total. Please contact Ecom right away.', []
    
    for sku, orderInfo in orders.items():
        if sku in inventoryMaster:
            if int(orderInfo['totalQty']) > inventoryMaster[sku]:
                outOfStockSKUs.append(sku)
                isOutOfStock = True
        else:
            cannotFindStockSKUs.append(sku)
    
    if isOutOfStock:
        return False, 'One or more items are out of stock.', outOfStockSKUs
        
    return True, '', []

def splitSKUs(orders):
    newOrders = {}

    for sku, orderInfo in orders.items():
        if '+' in sku:
            skus = sku.split('+')
            pricePerSku = orderInfo['pricePerPiece'] / len(skus)
            for newSku in skus:
                newOrderInfo = {
                    'sku': newSku,
                    'desc': orderInfo['desc'],
                    'totalQty': orderInfo['totalQty'],
                    'totalPrice': orderInfo['totalPrice'],
                    'pricePerPiece': pricePerSku
                }
                newOrders[newSku] = newOrderInfo
        else:
            newOrders[sku] = orderInfo

    return newOrders

def isTolerableOrderAmountDiscrepancy(totalOrderBeforeDiscount, grandTotalCrossCheck):
    if abs(totalOrderBeforeDiscount - grandTotalCrossCheck) < 0.1:
        return True
    return False

def processResult(filepath, uomMaster, inventoryMaster, orders, orderDetails):
    results = []
    grandTotalCrossCheck = 0
    invoiceTotal = 0

    orders = splitSKUs(orders)
    
    for sku, orderInfo in orders.items():
        caseUOM = uomMaster[sku + '-CASE']['uom'] if uomMaster.get(sku + '-CASE') else 0
        boxUOM = uomMaster[sku + '-BOX']['uom'] if uomMaster.get(sku +'-BOX') else 0
        grandTotalCrossCheck += getOrdersWithUOMVariants(results, orderInfo, caseUOM, boxUOM)

    if orderDetails['totalPaidByCustomer'] > 0:
        invoiceTotal = orderDetails['totalPaidByCustomer'] - orderDetails['totalOrderTax'] - orderDetails['totalShipping']
    else:
        invoiceTotal = orderDetails['totalOrderAmount'] - orderDetails['totalOrderTax'] - orderDetails['totalShipping']
    
    totalOrderBeforeDiscount = orderDetails['totalOrderAmount'] - orderDetails['totalOrderTax'] - orderDetails['totalShipping']
    # if isTolerableOrderAmountDiscrepancy(totalOrderBeforeDiscount, grandTotalCrossCheck):
    #     totalOrderBeforeDiscount = grandTotalCrossCheck

    discount = abs(invoiceTotal - totalOrderBeforeDiscount)

    resultDetails = {
        'grandTotal': grandTotalCrossCheck,
        'totalOrderBeforeDiscount': totalOrderBeforeDiscount,
        'totalOrderAmount': orderDetails['totalOrderAmount'],
        'totalPaidByCustomer': orderDetails['totalPaidByCustomer'],
        'totalTax': orderDetails['totalOrderTax'],
        'totalShipping': orderDetails['totalShipping'],
        'discount': discount,
        'invoiceTotal': totalOrderBeforeDiscount - discount
    }

    results.sort(key=cmp_to_key(sortOrders))

    dataFrame = pd.DataFrame(results, columns=['SKU', 'Desc', 'UOM', 'QTY', 'PpP'])
    dataFrame.index = dataFrame.index + 1
    # dataFrame.to_excel(filepath, startrow=3, startcol=0) # WORKING

    with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
        dataFrame.to_excel(writer, startrow=10, startcol=0)

        worksheet = writer.sheets['Sheet1']
        worksheet.write(0, 0, 'Grand Total: ${:.2f}'.format(resultDetails['grandTotal']))
        worksheet.write(1, 0, 'Order Total Before Discount: ${:.2f}'.format(resultDetails['totalOrderBeforeDiscount']))
        worksheet.write(2, 0, 'Order Total: ${:.2f}'.format(resultDetails['totalOrderAmount']))
        worksheet.write(3, 0, 'Customer Paid: ${:.2f}'.format(resultDetails['totalPaidByCustomer']))
        worksheet.write(4, 0, 'Tax: ${:.2f}'.format(resultDetails['totalTax']))
        worksheet.write(5, 0, 'Discount: ${:.2f}'.format(resultDetails['discount']))
        worksheet.write(6, 0, 'Shipping: ${:.2f}'.format(resultDetails['totalShipping']))
        worksheet.write(7, 0, 'Invoice Total: ${:.2f}'.format(resultDetails['invoiceTotal']))

    print('Your batch output file is: ' + filepath)
    
    isValidated, validationMessage, outOfStockSKUs = resultIsValidated(resultDetails, orders, inventoryMaster)
    if not isValidated:
        return False, validationMessage, outOfStockSKUs

    return True, '', []

def validateInputFilename(filename):
    cleaned = filename
    if '/' in filename:
        cleaned = filename.split('/')[-1]

    if '.csv' not in cleaned:
        cleaned = cleaned + '.csv'

    return USER_DOWNLOADS + cleaned

def getUOMMasterFilepath():
    return os.path.join(ASSETS_BASE_DIR, UOM_MASTER_FILENAME)

def getInventoryMasterFilepath():
    return os.path.join(ASSETS_BASE_DIR, INVENTORY_MASTER_FILENAME)

def writeLog(timestamp, status):
    path = os.path.join(ASSETS_BASE_DIR, LOGS_FILENAME)
    user = os.getenv('COMPUTERNAME')

    items = status["notExistSKUs"] or status["outOfStockSKUs"]

    try:
        with open(path, 'a') as file:
            file.write('USR;{} | IN;{} | SUCCESS;{} | ERR;{} | WARNING;{} | WARN;{} | ITEMS;{} | OUT;{} | VER;{} | TS;{}\n'.format(user, status["inputFilename"], status["success"], status["errorMessage"], status["warning"], status["warningMessage"], items, status["outputFilename"], APP_VERSION, timestamp))
    except:
        print('*** Error: Failed to write to logs. ***')

def batchOrders(inputFilename):
    timestamp = getTimestamp()
    batchFilename = validateInputFilename(inputFilename)

    isSuccess = True
    errorMessage = ''
    response = {
        'success': isSuccess,
        'errorMessage': errorMessage,
        'notExistSKUs': [],
        'warning': None,
        'warningMessage': None,
        'outOfStockSKUs': [],
        'inputFilename': batchFilename,
        'outputFilename': ''
    }

    uomMasterFilepath = getUOMMasterFilepath()
    inventoryMasterFilepath = getInventoryMasterFilepath()

    uomMaster, uomMsg = getUOMMasterData(uomMasterFilepath)
    inventoryMaster, invMsg = getInventoryMasterData(inventoryMasterFilepath)
    print(uomMsg)
    print(invMsg)

    if not uomMaster:
        errorMessage = "Please check UOM Master: " + uomMasterFilepath
        isSuccess = False
        response["success"] = isSuccess
        response["errorMessage"] = errorMessage
        writeLog(timestamp, response)
        return response

    if not inventoryMaster:
        errorMessage = "Please check Inventory Master: " + inventoryMasterFilepath
        isSuccess = False
        response["success"] = isSuccess
        response["errorMessage"] = errorMessage
        writeLog(timestamp, response)
        return response

    ordersStatus, orders, orderMessage = getOrdersFromInputfile(batchFilename, uomMaster)

    if ordersStatus == -1:
        errorMessage = orderMessage
        isSuccess = False
        response["success"] = isSuccess
        response["errorMessage"] = errorMessage
        writeLog(timestamp, response)
        return response
    if ordersStatus == 0:
        errorMessage = orderMessage
        isSuccess = False
        response["success"] = isSuccess
        response["errorMessage"] = errorMessage
        response["notExistSKUs"] = orders
        writeLog(timestamp, response)
        return response

    combinedOrders, orderDetails = combineOrders(orders, uomMaster)

    if not combinedOrders or not orderDetails:
        errorMessage = "Please try again."
        isSuccess = False
        response["success"] = isSuccess
        response["errorMessage"] = errorMessage
        writeLog(timestamp, response)
        return response

    outputFilename = 'batch_output_{}.xlsx'.format(timestamp)
    outputFilepath = OUTPUT_DIR + outputFilename

    isWithoutWarning, warningMessage, outOfStockSKUs =  processResult(outputFilepath, uomMaster, inventoryMaster, combinedOrders, orderDetails)
    response["warning"] = isWithoutWarning
    response["warningMessage"] = warningMessage
    response["outOfStockSKUs"] = outOfStockSKUs
    response["outputFilename"] = outputFilepath

    writeLog(timestamp, response)
    return response